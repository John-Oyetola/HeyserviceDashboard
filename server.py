#!/usr/bin/env python3
"""
Hey Service - Secure Authentication Server
Handles login, registration, and password reset via login details.xlsx
Uses SHA-256 + salt for secure password hashing.
Run: python3 server.py
Access: http://localhost:8000
"""

import http.server
import json
import os
import hashlib
import secrets
import urllib.parse
from pathlib import Path

import openpyxl

# === Configuration ===
# For Render or Heroku, they pass the port via an environment variable. We default to 8000 for localhost.
PORT = int(os.environ.get("PORT", 8000))
XLSX_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'login details.xlsx')

# === Password Security Functions ===

def hash_password(password):
    """Hash a password with a random salt using SHA-256."""
    salt = secrets.token_hex(16)
    hashed = hashlib.sha256((salt + password).encode('utf-8')).hexdigest()
    return f"{salt}:{hashed}"

def verify_password(password, stored_hash):
    """Verify a password against a stored salt:hash string."""
    if ':' not in stored_hash:
        # Legacy plain-text password — verify and return True to trigger migration
        return password == stored_hash
    salt, hashed = stored_hash.split(':', 1)
    check = hashlib.sha256((salt + password).encode('utf-8')).hexdigest()
    return check == hashed

def is_hashed(value):
    """Check if a password value is already hashed (contains salt:hash format)."""
    return ':' in value and len(value) > 40

# === Excel Read/Write ===

def read_users():
    """Read all users from the Excel file. Returns list of dicts."""
    if not os.path.exists(XLSX_FILE):
        # Create file with headers if it doesn't exist
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["username", "password", "role", "email", "company", "assigned_brand"])
        wb.save(XLSX_FILE)

    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active
    users = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if row_idx == 1:
            # Check if first row is headers
            first_cell = str(row[0] or '').strip().lower()
            if first_cell in ('username', 'user', 'name'):
                continue  # Skip header row

        username = str(row[0] or '').strip()
        password = str(row[1] or '').strip()
        role = str(row[2] or 'user').strip().lower() if len(row) > 2 and row[2] else 'user'
        email = str(row[3] or '').strip() if len(row) > 3 and row[3] else ''
        company = str(row[4] or '').strip() if len(row) > 4 and row[4] else ''
        assigned_brand = str(row[5] or 'all').strip() if len(row) > 5 and row[5] else 'all'

        if username:
            users.append({
                'username': username,
                'password': password,
                'role': role,
                'email': email,
                'company': company,
                'assigned_brand': assigned_brand,
                'row': row_idx
            })
    
    wb.close()
    return users

def add_user(username, password_hash, role, email, company, assigned_brand='all'):
    """Append a new user row to the Excel file."""
    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active
    ws.append([username, password_hash, role, email, company, assigned_brand])
    wb.save(XLSX_FILE)
    wb.close()

def update_password(row_idx, new_password_hash):
    """Update the password cell for a specific row."""
    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active
    ws.cell(row=row_idx, column=2, value=new_password_hash)
    wb.save(XLSX_FILE)
    wb.close()

def migrate_plaintext_password(row_idx, plain_password):
    """Migrate a plain-text password to a hashed version."""
    hashed = hash_password(plain_password)
    update_password(row_idx, hashed)
    return hashed

def username_exists(username):
    """Check if a username already exists (case-insensitive)."""
    users = read_users()
    return any(u['username'].lower() == username.lower() for u in users)

# === HTTP Request Handler ===

class HeyServiceHandler(http.server.SimpleHTTPRequestHandler):
    """Custom handler that serves static files AND handles API endpoints."""

    def do_OPTIONS(self):
        """Handle CORS preflight requests."""
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_POST(self):
        """Handle POST requests for API endpoints."""
        parsed = urllib.parse.urlparse(self.path)
        
        if parsed.path == '/api/login':
            self.handle_login()
        elif parsed.path == '/api/register':
            self.handle_register()
        elif parsed.path == '/api/forgot-password':
            self.handle_forgot_password()
        else:
            self.send_json_response(404, {'error': 'Endpoint not found'})

    def read_body(self):
        """Read and parse JSON body from request."""
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length)
        return json.loads(body.decode('utf-8'))

    def send_json_response(self, status, data):
        """Send a JSON response."""
        self.send_response(status)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps(data).encode('utf-8'))

    def handle_login(self):
        """POST /api/login — Verify credentials securely."""
        try:
            data = self.read_body()
            username = data.get('username', '').strip().lower()
            password = data.get('password', '').strip()

            if not username or not password:
                self.send_json_response(400, {'error': 'Username and password are required.'})
                return

            users = read_users()
            found = None
            for u in users:
                if u['username'].lower() == username:
                    found = u
                    break

            if not found:
                self.send_json_response(401, {'error': 'Invalid username or password.'})
                return

            # Verify password
            if verify_password(password, found['password']):
                # If password was plain-text, migrate it to hashed version
                if not is_hashed(found['password']):
                    migrate_plaintext_password(found['row'], password)
                    print(f"[SECURITY] Migrated plain-text password for user: {username}")

                self.send_json_response(200, {
                    'success': True,
                    'role': found['role'],
                    'username': found['username'],
                    'assigned_brand': found['assigned_brand'],
                    'company': found['company'],
                    'message': 'Login successful.'
                })
            else:
                self.send_json_response(401, {'error': 'Invalid username or password.'})

        except Exception as e:
            print(f"[ERROR] Login: {e}")
            self.send_json_response(500, {'error': 'Server error during authentication.'})

    def handle_register(self):
        """POST /api/register — Create new account and write to Excel."""
        try:
            data = self.read_body()
            username = data.get('username', '').strip()
            password = data.get('password', '').strip()
            email = data.get('email', '').strip()
            company = data.get('company', '').strip()

            if not username or not password:
                self.send_json_response(400, {'error': 'Username and password are required.'})
                return

            if len(password) < 6:
                self.send_json_response(400, {'error': 'Password must be at least 6 characters.'})
                return

            if username_exists(username):
                self.send_json_response(409, {'error': 'Username already exists. Please choose a different one.'})
                return

            # Hash the password securely
            password_hash = hash_password(password)

            # Write to Excel, passing company as the assigned_brand so they only see their own company
            add_user(username, password_hash, 'user', email, company, company)

            print(f"[REGISTER] New user created: {username} ({email}) - Company: {company}")

            self.send_json_response(201, {
                'success': True,
                'message': 'Account created successfully! You can now log in.'
            })

        except Exception as e:
            print(f"[ERROR] Register: {e}")
            self.send_json_response(500, {'error': 'Server error during registration.'})

    def handle_forgot_password(self):
        """POST /api/forgot-password — Reset password in Excel."""
        try:
            data = self.read_body()
            username = data.get('username', '').strip().lower()
            email = data.get('email', '').strip().lower()
            new_password = data.get('newPassword', '').strip()

            if not username or not email or not new_password:
                self.send_json_response(400, {'error': 'All fields are required.'})
                return

            if len(new_password) < 6:
                self.send_json_response(400, {'error': 'New password must be at least 6 characters.'})
                return

            users = read_users()
            found = None
            for u in users:
                if u['username'].lower() == username:
                    found = u
                    break

            if not found:
                self.send_json_response(404, {'error': 'Username not found.'})
                return

            # Verify email matches (if stored)
            if found['email'] and found['email'].lower() != email:
                self.send_json_response(403, {'error': 'Email does not match the account on file.'})
                return

            # Hash the new password and update Excel
            new_hash = hash_password(new_password)
            update_password(found['row'], new_hash)

            print(f"[PASSWORD RESET] Password updated for user: {username}")

            self.send_json_response(200, {
                'success': True,
                'message': 'Password has been reset successfully! You can now log in with your new password.'
            })

        except Exception as e:
            print(f"[ERROR] Forgot Password: {e}")
            self.send_json_response(500, {'error': 'Server error during password reset.'})

    def log_message(self, format, *args):
        """Custom log formatting."""
        if '/api/' in str(args[0]) if args else False:
            print(f"[API] {args[0]}")
        else:
            super().log_message(format, *args)

# === Start Server ===

if __name__ == '__main__':
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    
    print("=" * 50)
    print("  Hey Service - Secure Authentication Server")
    print("=" * 50)
    print(f"  Serving on: http://localhost:{PORT}")
    print(f"  Excel DB:   {XLSX_FILE}")
    print(f"  Security:   SHA-256 + Salt Password Hashing")
    print("=" * 50)
    
    server = http.server.HTTPServer(('', PORT), HeyServiceHandler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[SERVER] Shutting down...")
        server.shutdown()
